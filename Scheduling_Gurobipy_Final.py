from numpy import random
import string
import gurobipy as gp
from gurobipy import GRB
from openpyxl import Workbook

number_machines = 6
number_jobs = 10
iterations = 100
M = 100

file_path = 'C:/Users/oguzt/PycharmProjects/shahin_new_results/'+str(number_machines) + 'Machines_' + str(number_jobs) + 'Jobs.xlsx'
file_path2 = 'C:/Users/oguzt/PycharmProjects/shahin_new_results/'+str(number_machines) + 'Machines_' + str(number_jobs) + 'Jobs_models.xlsx'
wb = Workbook()
wb_models = Workbook()
wb.worksheets[0].append(list(range(20, -1, -1)))
wb.save(file_path)
wb_models.save(file_path2)

def generator(number_machines, number_jobs):
    jobs = list(range(number_jobs))
    machines = []
    letters = list(string.ascii_letters)
    for mach in range(number_machines):
        poper = letters.pop(0)
        machines.append(poper)
    orders = [x+1 for x in list(range(len(machines)))]
    order_job_set = []
    work_order = {}
    machine_job_set = []
    process_time = {}
    model_rec = {}
    for job in jobs:
        temp_machines = list(machines)
        for order in orders:
            random.shuffle(temp_machines)
            node = tuple([order, job])
            order_job_set.append(node)
            work_order[node] = temp_machines.pop(0)
            model_rec[node] = work_order[node]
            node_p = tuple([machines[(order - 1)], job])
            machine_job_set.append(node_p)
            process_time[node_p] = random.randint(1, 10)
            model_rec[node] = [model_rec[node], process_time[node_p]]
    return jobs, machines, orders, order_job_set, machine_job_set, work_order, process_time, model_rec


def model_builder(number_machines, number_jobs):
    jobs, machines, orders, order_job_set, machine_job_set, work_order, process_time, model_rec = generator(number_machines, number_jobs)
    model_data = [jobs, machines, orders, order_job_set, machine_job_set, work_order, process_time, model_rec]
    m = gp.Model()
    t = m.addVars(machines, jobs, vtype=GRB.INTEGER, name="t")
    s = m.addVars(machines, jobs, vtype=GRB.INTEGER, name="s")
    y = m.addVars(machines, jobs, jobs, vtype=GRB.BINARY, name="y")
    objectemun = sum(t[work_order[i, j], j] for i in orders for j in jobs)
    m.setObjective(sense=gp.GRB.MINIMIZE, expr=objectemun)
    StartFinish = m.addConstrs((s[i, j] == (t[i, j] - process_time[i, j]) for i in machines for j in jobs), name='StartFinish')
    StartAfterFinish = m.addConstrs((s[work_order[k, j], j] - 1 >= t[work_order[k-1, j], j] for k in (orders[1:]) for j in jobs), name='StartAfterFinish')
    for i in machines:
        for j in jobs:
            for k in jobs:
                if k != j:
                    m.addConstr(s[i, k] >= t[i, j] - y[i, j, k] * M)
                    m.addConstr(s[i, j] >= t[i, k] - (1 - y[i, j, k]) * M)
    m.update()
    return m, model_rec


def my_callback(m, cb_where):
    if cb_where == GRB.Callback.MIP:
        objbst = m.cbGet(GRB.Callback.MIP_OBJBST)
        objbnd = m.cbGet(GRB.Callback.MIP_OBJBND)
        gap = abs((objbst - objbnd) / objbst)*100
        runtime = m.cbGet(GRB.Callback.RUNTIME)
        global time_gap
        global times
        if gap < 20 and time_gap == 0:
            print('GAP ::::', 20, '%', '::::', round(runtime, 5))
            times[20] = round(runtime, 5)
            time_gap = 1
        elif gap < 19 and time_gap == 1:
            print('GAP ::::', 19, '%', '::::', round(runtime, 5))
            times[19] = round(runtime, 5)
            time_gap = 2
        elif gap < 18 and time_gap == 2:
            print('GAP ::::', 18, '%', '::::', round(runtime, 5))
            times[18] = round(runtime, 5)
            time_gap = 3
        elif gap < 17 and time_gap == 3:
            print('GAP ::::', 17, '%', '::::', round(runtime, 5))
            times[17] = round(runtime, 5)
            time_gap = 4
        elif gap < 16 and time_gap == 4:
            print('GAP ::::', 16, '%', '::::', round(runtime, 5))
            times[16] = round(runtime, 5)
            time_gap = 5
        elif gap < 15 and time_gap == 5:
            print('GAP ::::', 15, '%', '::::', round(runtime, 5))
            times[15] = round(runtime, 5)
            time_gap = 6
        elif gap < 14 and time_gap == 6:
            print('GAP ::::', 14, '%', '::::', round(runtime, 5))
            times[14] = round(runtime, 5)
            time_gap = 7
        elif gap < 13 and time_gap == 7:
            print('GAP ::::', 13, '%', '::::', round(runtime, 5))
            times[13] = round(runtime, 5)
            time_gap = 8
        elif gap < 12 and time_gap == 8:
            print('GAP ::::', 12, '%', '::::', round(runtime, 5))
            times[12] = round(runtime, 5)
            time_gap = 9
        elif gap < 11 and time_gap == 9:
            print('GAP ::::', 11, '%', '::::', round(runtime, 5))
            times[11] = round(runtime, 5)
            time_gap = 10
        elif gap < 10 and time_gap == 10:
            print('GAP ::::', 10, '%', '::::', round(runtime, 5))
            times[10] = round(runtime, 5)
            time_gap = 11
        elif gap < 9 and time_gap == 11:
            print('GAP ::::', 9, '%', '::::', round(runtime, 5))
            times[9] = round(runtime, 5)
            time_gap = 12
        elif gap < 8 and time_gap == 12:
            print('GAP ::::', 8, '%', '::::', round(runtime, 5))
            times[8] = round(runtime, 5)
            time_gap = 13
        elif gap < 7 and time_gap == 13:
            print('GAP ::::', 7, '%', '::::', round(runtime, 5))
            times[7] = round(runtime, 5)
            time_gap = 14
        elif gap < 6 and time_gap == 14:
            print('GAP ::::', 6, '%', '::::', round(runtime, 5))
            times[6] = round(runtime, 5)
            time_gap = 15
        elif gap < 5 and time_gap == 15:
            print('GAP ::::', 5, '%', '::::', round(runtime, 5))
            times[5] = round(runtime, 5)
            time_gap = 16
        elif gap < 4 and time_gap == 16:
            print('GAP ::::', 4, '%', '::::', round(runtime, 5))
            times[4] = round(runtime, 5)
            time_gap = 17
        elif gap < 3 and time_gap == 17:
            print('GAP ::::', 3, '%', '::::', round(runtime, 5))
            times[3] = round(runtime, 5)
            time_gap = 18
        elif gap < 2 and time_gap == 18:
            print('GAP ::::', 2, '%', '::::', round(runtime, 5))
            times[2] = round(runtime, 5)
            time_gap = 19
        elif gap < 1 and time_gap == 19:
            print('GAP ::::', 1, '%', '::::', round(runtime, 5))
            times[1] = round(runtime, 5)
            time_gap = 20


for itr in range(iterations):
    times = {x: 0 for x in range(20, -1, -1)}
    time_gap = 0
    time_gap_plus = 20
    print(itr, 'Iteration--------------------------------')
    m, model_rec = model_builder(number_machines, number_jobs)
    m.setParam('OutputFlag', False)
    m.setParam('Threads', 24)
    m.optimize(my_callback)
    times[0] = round(m.Runtime, 5)
    print('GAP ::::', 0, '%', '::::', round(m.Runtime, 5))
    m.terminate()
    wb.worksheets[0].append(list(times.values()))
    list_ = [str(i) for i in list(model_rec.keys())]
    wb_models.worksheets[0].append(list_)
    wb_models.save(file_path2)
    column_ = 1
    for key, value in model_rec.items():
        cell_to_write = wb_models.worksheets[0].cell(row=itr+2, column=column_)
        cell_to_write.value = str(value)
        column_ += 1
    wb_models.save(file_path2)
    wb.save(file_path)
    time_gap = 0

